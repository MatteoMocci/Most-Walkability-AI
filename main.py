import sardegna_scripts as sard
import itertools
import time
import os

from sklearn.feature_selection import SelectFromModel
from sklearn.svm import SVC, LinearSVC
from sklearn.preprocessing import RobustScaler
import matplotlib.pyplot as plt

from transformers import AutoModelForImageClassification, AutoFeatureExtractor
import torch
from torch.utils.data import DataLoader
import numpy as np
from tqdm import tqdm
import torch.nn as nn
from transformers import pipeline, AutoImageProcessor
import pandas as pd
import joblib
from torchvision import transforms
import dual_encoder as de
from transformers import TrainingArguments, Trainer
from paired_datasets.paired_image_dataset import StreetSatDataset
import paired_datasets.paired_image_dataset as paired
import shutil
from openpyxl import load_workbook


'''
Questa funzione permette di testare diverse configurazioni di modelli specificando una lista di checkpoint, di epoche, di learning rate, di batch_size e di ottimizzatori
Viene generata una lista con tutte le possibili combinazioni di questi valori e per ciascuna combinazione viene addestrato un modello, calcolate le metriche sul test set e i dati vengono salvati su un file excel
'''
def test_configurations():

    checkpoints = ['google/vit-base-patch16-224','timm/efficientnet_b3.ra2_in1k',
                   'timm/resnet18.a1_in1k','microsoft/beit-base-patch16-224-pt22k-ft22k','timm/resnet50.a1_in1k']
    n_epochs = [10,15]
    lrs = [1e-4]
    batch_sizes = [16,32]
    optim = ['adamw_hf', 'adafactor']

    hyperparam_combinations = list(itertools.product(checkpoints, n_epochs, lrs, batch_sizes, optim))

    # Experiment loop
    for idx, (checkpoint, n_epoch, lr, batch_size, optimizer) in enumerate(hyperparam_combinations, start=2):
        print(f"Config num {idx-1}/{len(hyperparam_combinations)}")
        start = time.time()
        metrics = sard.training_step(n=n_epoch,lr=lr,opt=optimizer,batch_size=batch_size)
        end = time.time()
        elapsed = end - start
        sard.save_results_to_excel(idx=idx,checkpoint=checkpoint,n_epoch=n_epoch,lr=lr,optimizer=optimizer,batch_size=batch_size,metrics=metrics,elapsed=elapsed,name='best_results.xlsx')

'''
Questa funzione permette l'estrazione delle feature per il modello model sui dati contenuti in dataloader. Queste feature vengono salvate in "save_path"
@param model        il modello pre-addestrato che verrà usato per l'estrzione degli embedding
@param dataloader   il dataloader contenente i batch delle immagini da cui estrarre le feature
@param device       cpu o, se CUDA è disponibile, gpu
@param save_path    il path in cui salvare le feature estratte
@return le feature estratte
'''
def extract_features(model, dataloader, device, save_path):
    model.eval()
    features = []
    with torch.no_grad():
        for batch in tqdm(dataloader, desc=f"Extracting features for {save_path}"):
            inputs = batch['pixel_values'].to(device)
            outputs = model(inputs, output_hidden_states=True)
            feature = outputs.hidden_states[-1][:, 0, :]  # [batch_size, sequence_length, hidden_size] -> [batch_size, hidden_size]
            features.append(feature.cpu().numpy())
            # Print the shape of the embeddings from the hidden layers
    features = np.concatenate(features, axis=0)
    np.save(save_path, features)
    return features

def extract_inference_features(model, images, device, save_path):
    model.eval()
    features = []
    
    # Convert images to tensors
    processor = AutoImageProcessor.from_pretrained("google/vit-base-patch16-224", ignore_mismatched_sizes=True)
    inputs = processor(images, return_tensors='pt')['pixel_values'].to(device)
    
    with torch.no_grad():
        for i in tqdm(range(0, len(inputs)), desc=f"Extracting features for {save_path}"):
            batch = inputs[i:i+1]  # Process one image at a time (could batch more if needed)
            outputs = model(batch, output_hidden_states=True)
            feature = outputs.hidden_states[-1][:, 0, :]  # Extract embeddings from the hidden state
            features.append(feature.cpu().numpy())
    
    features = np.concatenate(features, axis=0)
    np.save(save_path, features)
    return features
'''
Questa funzione effettua il fit di un feature selector basato su Random Forest sulle feature di training.
Questo feature selector viene poi usato per effettuare feature selection sulle feature di training
@param X            le feature estratte dal modello sui dati di training
@param y            le label delle istanze di training
@param save_path    il path in cui salvare le feature selezionate
@return le feature di training selezionate e il feature selector da usare poi su feature di validation e test
'''
def fit_feature_selector(X, y, save_path):
    rf = LinearSVC(C=0.01, penalty="l2", dual=False).fit(X, y)
    selector = SelectFromModel(rf, prefit=True)
    X_new = selector.transform(X)
    print(f"Prima: {X.shape} \t Dopo:{X_new.shape}")

    # Extract feature coefficients
    coefficients = np.mean(np.abs(rf.coef_), axis=0)
    
    # Define the bins for feature importance
    bins = np.linspace(0, max(coefficients), 10)
    
    # Count the number of features in each bin
    hist, bin_edges = np.histogram(coefficients, bins=bins)

    # Get the threshold used by the feature selector
    threshold = np.mean(np.abs(rf.coef_))
    print(f"Feature selection threshold: {threshold}")

    to_plot = False
    # Plot the histogram of feature importance
    if to_plot:
        plt.figure(figsize=(10, 6))
        plt.bar(bin_edges[:-1], hist, width=np.diff(bin_edges), edgecolor="black", align="edge")
        plt.axvline(threshold, color='red', linestyle='--', linewidth=2, label='Feature Selection Threshold')
        plt.xlabel('Feature Importance Range')
        plt.ylabel('Number of Features')
        plt.title('Distribution of Feature Importances')
        # plt.show()

    np.save(save_path, X_new)
    return X_new, selector

'''
Questa funzione permette di stampare a video un istogramma che rappresenta la distribuzione delle feature selezionate. 
Nell'asse delle x è presente il valore di importanza, discretizzato in 10 range e nell'asse delle y il numero di feature che ha tale importanza.
Ogni barra dell'istogramma è divisa in 2: una parte blu che indica il numero di feature appartenenti al modello streetview e una arancione per il numero di feature appartenenti al modello satellitare
@param X                Le feature di training combinate
@param y                Le label delle istanze di training
@param num_features1    Il numero di features estratte dal primo modello, per poter dividere le feature concatenate in base al modello da cui sono esrtatte.
'''
def print_combined_feature_importance(X, y, num_features1):
    rf = LinearSVC(C=0.01, penalty="l2", dual=False).fit(X, y)
    # Extract feature coefficients
    coefficients = np.mean(np.abs(rf.coef_), axis=0)

    # Define the bins for feature importance
    bins = np.linspace(0, max(coefficients), 10)
    
    # Count the number of features in each bin for both models
    hist1, _ = np.histogram(coefficients[:num_features1], bins=bins)
    hist2, _ = np.histogram(coefficients[num_features1:], bins=bins)

    # Plot the stacked bar chart
    to_plot = False
    if to_plot:
        plt.figure(figsize=(10, 6))
        plt.bar(bins[:-1], hist1, width=np.diff(bins), edgecolor="black", align="edge", label='Feature Modello Street-view', color='blue')
        plt.bar(bins[:-1], hist2, width=np.diff(bins), edgecolor="black", align="edge", bottom=hist1, label='Feature Modello Satellitare', color='orange')
        plt.xlabel('Feature Importance Range')
        plt.ylabel('Number of Features')
        plt.title('Distribution of Feature Importances')
        plt.legend()
        plt.show()

'''
Questa funzione effettua il passaggio di feature selection per feature di validation e test.
@param selector     il feature selector fittato sulle feature di training tramite "fit_feature_selector()"
@param X            le feature estratte dal modello sui dati di validation o test
@param save_path    il path in cui salvare le feature selezionate
return le feature di X selezionate tramite il feature selector
'''
def select_features(selector, X, save_path):
    selected = selector.transform(X)
    np.save(save_path, selected)
    return selected

'''
Questa funzione effettua la concatenazione di due array numpy di feature X1 e X2, salva e restituisce un unico array numpy X_combined
@param  X1          il vettore di feature estratte dal train, validation o test delle immagini street-view
@param  X2          il vettore di feature estratte dal train, validation o test delle immagini satellitari
@param  save_path   il path dove salvare le feature concatenate
@return le feature di immagini street-view e satellitari concatenate
'''
def concatenate_features(X1, X2, save_path):
    # I due vettori devono avere lo stesso numero di righe per essere concatenati
    if X1.shape[0] != X2.shape[0]:
        raise ValueError("The number of samples in both datasets must be the same.")
    
    # Concatena lungo l'asse delle feature (axis=1)
    X_combined = np.concatenate((X1, X2), axis=1)
    
    # Salva le feature concatenate
    np.save(save_path, X_combined)
    
    return X_combined

import torch
import torch.nn as nn

class CombinedModel(nn.Module):
    def __init__(self, input_shape, num_classes, pooling='mean'):
        """
        Parameters:
          input_shape (tuple): A tuple (dim1, dim2) describing the shape of the input
                               tensor per sample (excluding batch dimension).
          num_classes (int):   Number of output classes.
          pooling (str):       Aggregation method over the dim1 dimension. Options are:
                               'mean', 'max', or 'flatten'. 
        """
        super(CombinedModel, self).__init__()
        self.pooling = pooling.lower()
        self.input_shape = input_shape  # e.g. (98, 768) or any (dim1, dim2)

        if self.pooling in ['mean', 'max']:
            # After pooling over dim1, the feature vector will have dimension equal to dim2.
            fc_input_dim = input_shape[1]
        elif self.pooling == 'flatten':
            # Flattening retains all features: dim1 * dim2.
            fc_input_dim = input_shape[0] * input_shape[1]
        else:
            raise ValueError(f"Unsupported pooling method: {pooling}. "
                             "Choose one of 'mean', 'max', or 'flatten'.")

        self.classifier = nn.Sequential(
            nn.Linear(fc_input_dim, 512),
            nn.ReLU(),
            nn.Dropout(0.3),
            nn.Linear(512, num_classes)
        )

    def forward(self, x):
        """
        Forward pass.
        
        Args:
          x (torch.Tensor): A tensor of shape [batch, dim1, dim2]
        Returns:
          torch.Tensor: The output logits of shape [batch, num_classes]
        """
        if self.pooling == 'mean':
            # Average along the dim1 dimension
            x = x.mean(dim=1)  # Resulting shape: [batch, dim2]
        elif self.pooling == 'max':
            # Maximum along the dim1 dimension
            x, _ = x.max(dim=1)  # Resulting shape: [batch, dim2]
        elif self.pooling == 'flatten':
            # Flatten the [dim1, dim2] into one dimension
            x = x.view(x.size(0), -1)  # Resulting shape: [batch, dim1 * dim2]
        return self.classifier(x)

'''
Questo metodo effettua la classificazione finale delle feature raccolte da entrambe le immagini tramite Support Vector Machine con Kernel rbf.
@param train_combined           le feature selezionate di training per dati streetview e satellitari
@param valid_combined           le feature selezionate di validation per dati streetview e satellitari
@param test_combined            le feature selezionate di test per dati streetview e satellitari
@param train_street             le feature di training per i dati street-view (serve per ricavare le label)
@param valid_street             le feature di validation per i dati street-view (serve per ricavare le label)
@param test_street              le feature di test per i dati street-view (serve per ricavare le label)
@param train_street_selected    le feature selezionate dai dati streetview (serve per ricavare il numero di feature selezionate dal modello street-view)
@param train_sat_selected       le feature selezionate dai dati satellitari (serve per ricavare il numero di feature selezionate dal modello satellitare)
'''
def final_classification_svm(train_combined, valid_combined, test_combined, train_street, valid_street, test_street, train_street_selected, train_sat_selected, train_street_labels, valid_street_labels, test_street_labels):
    svm_model = SVC(kernel='rbf', degree=3, C=1.0, random_state=42, probability=True)                                                                              # Inizializza il modello SVM
    svm_model.fit(train_combined, train_street_labels)                                                                                                           # Fitta il modello
    joblib.dump(svm_model, "svm.pkl")


    print_combined_feature_importance(train_combined, train_street_labels , train_street_selected.shape[1])
    sard.test_svm_model(svm_model, valid_combined, test_combined, train_street, valid_street, test_street, train_street_labels=train_street_labels, valid_street_labels=valid_street_labels, test_street_labels=test_street_labels)                                                         # Validation e test del modello

def inference_svm(combined):
    svm_model = joblib.load("svm.pkl")
    return svm_model.predict(combined)
    

'''
Questo metodo effettua la classificazione finale delle feature raccolte da entrambe le immagini tramite una rete deep torch.
@param train_combined           le feature selezionate di training per dati streetview e satellitari
@param valid_combined           le feature selezionate di validation per dati streetview e satellitari
@param test_combined            le feature selezionate di test per dati streetview e satellitari
@param train_street             le feature di training per i dati street-view (serve per ricavare le label)
@param valid_street             le feature di validation per i dati street-view (serve per ricavare le label)
@param test_street              le feature di test per i dati street-view (serve per ricavare le label)
@param train_street_selected    le feature selezionate dai dati streetview (serve per ricavare il numero di feature selezionate dal modello street-view)
@param train_sat_selected       le feature selezionate dai dati satellitari (serve per ricavare il numero di feature selezionate dal modello satellitare)
@param device                   cpu o GPU (cuda)
'''
def final_classification_deep(train_combined, valid_combined, test_combined, train_street, valid_street, test_street, train_street_selected, train_sat_selected,device,train_street_labels):

    num_classes = 5                                                                                                             # Numero classi

   
    combined_model = CombinedModel((train_combined.shape[1], train_combined.shape[2]), num_classes).to(device)         # Creazione modello combinato e caricamento su GPU

    criterion = nn.CrossEntropyLoss()                                                                                           # Definizione loss e optimizer
    optimizer = torch.optim.Adam(combined_model.parameters(), lr=1e-4)

    # print_combined_feature_importance(train_combined, train_street_labels,train_street_selected.shape[1])
    train_combined = torch.as_tensor(train_combined).to(device)                                                                    # Trasformazione dati training in tensore e caricamento su GPU


    # Training loop
    num_epochs = 10
    for epoch in range(num_epochs):
        combined_model.train()
        optimizer.zero_grad()
        outputs = combined_model(train_combined)
        loss = criterion(outputs, torch.tensor(train_street['label']).to(device))
        loss.backward()
        optimizer.step()
        
        print(f'Epoch [{epoch + 1}/{num_epochs}], Loss: {loss.item():.4f}')

    torch.save(combined_model.state_dict(), 'combined_model.pth')                                                               # Salvataggio modello
    res_valid, test_valid = sard.test_torch_model(combined_model, valid_combined,test_combined,train_street,valid_street,test_street,device)            # Valuta le performance del modello su Validation e test set
    return res_valid, test_valid
'''
Questa funzione esegue tutto ciò che è richiesto per addestrare il modello di classificazione della camminabilità di punti basandosi su immagini delle strade street-view e satellitari:
    1) Addestra o carica il modello delle immagini street-view
    2) Addestra o carica il modello delle immagini satellitari
    3) Estrae o carica le feature per i dati di training, validation e test dai modelli ottenuti nello step 1 e 2
    4) Esegue il fitting di un SVM con kernel lineare e penalità norma L2 per le feature di training ottenute nello step 3, per entrambi i modelli o carica le feature selezionate già presenti, saltando lo step 5
    5) Esegui feature selection anche su validation e test con il selector fittato sul training
    6) Normalizza le feature con Robust Scaler
    7) Concatena le feature estratte dal modello street e dal modello satellitare per avere un unico vettore di feature per il training, uno per il validation e uno per il test o carica quelle già presenti
    8) Fitta un SVM sulle feature con kernel radial based di training ed effettua validation e test
'''
def walkability_pipeline(checkpoint):
    force_training = True
    batch_size = 16
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
                                                                                             # Se CUDA è presente, usa la GPU 
    street_dataset = StreetSatDataset(mode='street')
    street_dataset.download_and_prepare()
    street_dataset_splits = street_dataset.as_dataset()

    sat_dataset = StreetSatDataset(mode='sat')
    sat_dataset.download_and_prepare()
    sat_dataset_splits = street_dataset.as_dataset()

    street_tf = paired.create_transforms(mode="street",checkpoint=checkpoint)
    sat_tf = paired.create_transforms(mode="sat",checkpoint=checkpoint)

    train_street_labels = street_dataset_splits["train"]["label"]
    valid_street_labels = street_dataset_splits["validation"]["label"]
    test_street_labels = street_dataset_splits["test"]["label"]


    train_street = street_dataset_splits["train"].with_transform(street_tf)
    valid_street = street_dataset_splits["validation"].with_transform(street_tf)
    test_street = street_dataset_splits["test"].with_transform(street_tf)
    
    train_sat = sat_dataset_splits["train"].with_transform(sat_tf)
    valid_sat = sat_dataset_splits["validation"].with_transform(sat_tf)
    test_sat = sat_dataset_splits["test"].with_transform(sat_tf)                                                                                                  # Carica i dataset per training, validation e test per le immagini satellitari

    if os.path.exists('./sardegna-vit') and not force_training:                                                                                                                          # Se il modello addestrato sulle immagini street view è presente
        street_model = AutoModelForImageClassification.from_pretrained("sardegna-vit").to(device)                                                                 # caricalo

    else:                                                                                                                                                         # altrimenti
        trainer_street = sard.create_trainer(train_street,valid_street,n=10,lr=1e-4,optim='adamw_hf',output_dir='./sardegna-vit',checkpoint=checkpoint)                                 # crea il trainer
        trainer_street = sard.train_model(trainer_street)                                                                                                         # esegui l'addestramento su street-view
        metrics_street = sard.test_model(trainer_street, test_street)
        print(metrics_street)                                                                                                                                     # testa il modello
        street_model = AutoModelForImageClassification.from_pretrained("sardegna-vit").to(device)                                                                 # carica il modello appena addestrato
    
    street_extractor = AutoImageProcessor.from_pretrained(checkpoint)

    if os.path.exists('./satellite-vit') and not force_training:                                                                                                                         # Se il modello addestrato sulle immagini satellitari è presente
        sat_model = AutoModelForImageClassification.from_pretrained("satellite-vit").to(device)
                                                                           
    else:                                                                                                                                                         # altrimenti
        trainer_sat = sard.create_trainer(train_sat,valid_sat,n=10,lr=1e-4,optim='adamw_hf',output_dir='./satellite-vit',checkpoint=checkpoint)                                         # crea il trainer
        trainer_sat = sard.train_model(trainer_sat)                                                                                                               # esegui l'addestramento su immagini satellitari
        metrics_sat = sard.test_model(trainer_sat, test_sat)                                                                                                      # testa il modello
        print(metrics_sat)
        sat_model = AutoModelForImageClassification.from_pretrained("satellite-vit").to(device)                                             # carica il modello appena addestrato
    sat_extractor = AutoImageProcessor.from_pretrained(checkpoint)
    
    if os.path.exists('./features/train_street_features.npy') and os.path.exists('./features/train_sat_features.npy') and False:                                            # Se sono presenti le feature di training
        train_street_features = np.load('features/train_street_features.npy')                                                                                     # carica feature di training, validation e test per immagini street-view e satellitari
        valid_street_features = np.load('features/valid_street_features.npy')
        test_street_features  = np.load('features/test_street_features.npy')

        train_sat_features = np.load('features/train_sat_fesatures.npy')
        valid_sat_features = np.load('features/valid_sat_features.npy')
        test_sat_features  = np.load('features/test_sat_features.npy')    
    else:
        new_batch_size = 16                                                                                                                                                         # altrimenti
        train_street_loader = DataLoader(train_street, batch_size=new_batch_size, shuffle=False, collate_fn=sard.collate_fn)                                          # crea i data loader per i dataset
        valid_street_loader = DataLoader(valid_street, batch_size=new_batch_size, shuffle=False, collate_fn=sard.collate_fn)
        test_street_loader = DataLoader(test_street, batch_size=new_batch_size, shuffle=False, collate_fn=sard.collate_fn)

        train_sat_loader = DataLoader(train_sat, batch_size=new_batch_size, shuffle=False, collate_fn=sard.collate_fn)
        valid_sat_loader = DataLoader(valid_sat, batch_size=new_batch_size, shuffle=False, collate_fn=sard.collate_fn)
        test_sat_loader = DataLoader(test_sat, batch_size=new_batch_size, shuffle=False, collate_fn=sard.collate_fn)
        
        '''
        train_street_features = extract_features(model=street_model,dataloader=train_street_loader, device=device,save_path='features/train_street_features.npy') # estrai le feature di training, validation e test per immagini satellitari
        valid_street_features = extract_features(model=street_model,dataloader=valid_street_loader, device=device,save_path='features/valid_street_features.npy')
        test_street_features = extract_features(model=street_model,dataloader=test_street_loader, device=device,save_path='features/test_street_features.npy')

        train_sat_features = extract_features(model=sat_model,dataloader=train_sat_loader, device=device,save_path='features/train_sat_features.npy')
        valid_sat_features = extract_features(model=sat_model,dataloader=valid_sat_loader, device=device,save_path='features/valid_sat_features.npy')
        test_sat_features = extract_features(model=sat_model,dataloader=test_sat_loader, device=device,save_path='features/test_sat_features.npy')
        '''
 
    if os.path.exists('./features/train_street_selected.npy') and os.path.exists('./features/train_sat_selected.npy'):                                           # Se sono presenti le feature selezionate di training, caricale
        train_street_selected = np.load('features/train_street_selected.npy')
        valid_street_selected = np.load('features/valid_street_selected.npy')
        test_street_selected = np.load('features/test_street_selected.npy')

        train_sat_selected = np.load('features/train_sat_selected.npy')
        valid_sat_selected = np.load('features/valid_sat_selected.npy')
        test_sat_selected = np.load('features/test_sat_selected.npy')
    else:
        '''
        if train_street_features.ndim > 2:
            train_street_features = train_street_features.reshape(train_street_features.shape[0],-1)
            valid_street_features = valid_street_features.reshape(valid_street_features.shape[0],-1)
            test_street_features = test_street_features.reshape(test_street_features.shape[0],-1)
            train_sat_features = train_sat_features.reshape(train_sat_features.shape[0],-1)
            valid_sat_features = valid_sat_features.reshape(valid_sat_features.shape[0],-1)
            test_sat_features = test_sat_features.reshape(test_sat_features.shape[0],-1)
            '''
        '''
        train_street_selected, street_selector = fit_feature_selector(train_street_features,train_street_labels,'features/train_street_selected.npy')           # altrimenti fitta un feature selector per street e sat sui dati di training e seleziona feature di train, valid e test
        train_sat_selected, sat_selector = fit_feature_selector(train_sat_features, train_sat_labels,'features/train_sat_selected.npy')

        valid_street_selected = select_features(street_selector, valid_street_features, 'features/valid_street_selected.npy')
        test_street_selected = select_features(street_selector, test_street_features, 'features/test_street_selected.npy')

        valid_sat_selected = select_features(sat_selector, valid_sat_features, 'features/valid_sat_selected.npy')
        test_sat_selected = select_features(sat_selector, test_sat_features, 'features/test_sat_selected.npy')

    
        train_street_selected = train_street_features
        train_sat_selected = train_sat_features
        valid_street_selected = valid_street_features
        valid_sat_selected = valid_sat_features
        test_street_selected = test_street_features
        test_sat_selected = test_sat_features
        '''
        
        train_street_selected = self_attention_fs(street_model, train_street_loader,'features/train_street_selected.npy')
        valid_street_selected = self_attention_fs(street_model, valid_street_loader, 'features/valid_street_selected.npy')
        test_street_selected = self_attention_fs(street_model, test_street_loader, 'features/test_street_selected.npy')
        train_sat_selected = self_attention_fs(sat_model, train_sat_loader, 'features/train_sat_selected.npy')
        valid_sat_selected = self_attention_fs(sat_model, valid_sat_loader, 'features/valid_sat_selected.npy')
        test_sat_selected = self_attention_fs(sat_model, test_sat_loader, 'features/test_sat_selected.npy')
   

    if os.path.exists("features/train_combined.npy"):                                                                                                              # Se sono presenti le feature concatenate di training
        train_combined = np.load("features/train_combined.npy")                                                                                                    # caricale per train, valid e test
        valid_combined = np.load("features/valid_combined.npy")                                                                                                          
        test_combined  = np.load("features/test_combined.npy")
    else:
        '''
        scaler = RobustScaler()                                                                                                                                    # normalizza feature                                                                                                                                   
        train_street_selected = scaler.fit_transform(train_street_selected)
        valid_street_selected = scaler.fit_transform(valid_street_selected)
        test_street_selected = scaler.fit_transform(test_street_selected)

        train_sat_selected = scaler.fit_transform(train_sat_selected)
        valid_sat_selected = scaler.fit_transform(valid_sat_selected)
        test_sat_selected = scaler.fit_transform(test_sat_selected)
        '''

        train_combined = concatenate_features(train_street_selected, train_sat_selected, "features/train_combined.npy")                                            # altrimenti concatena usando np.concatenate per train, valid e test
        valid_combined = concatenate_features(valid_street_selected, valid_sat_selected, "features/valid_combined.npy")                                                    
        test_combined = concatenate_features(test_street_selected, test_sat_selected, "features/test_combined.npy")

    use_svm = False
    if use_svm:
        final_classification_svm(train_combined,valid_combined,test_combined,train_street,valid_street,test_street,train_street_selected,train_sat_selected,train_street_labels,valid_street_labels,test_street_labels)            # Effettua la classificazione finale con SVM
    else:
        res_valid, res_test = final_classification_deep(train_combined,valid_combined,test_combined,train_street,valid_street,test_street,train_street_selected,train_sat_selected,device,train_street_labels)    # Altrimenti usa una rete torch custom
    return res_valid, res_test
def data_inference():
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    batch_size = 16

    # Parte 1 - Immagini Streetview

    # Caricare le immagini Streetview
    street_valdala = sard.get_all_images_of_folder('C:/Users/mocci/Desktop/MOST/Val d\'Ala/dataset_merged')
    names = [os.path.splitext(os.path.basename(im.filename))[0] for im in street_valdala]
    osm_lat_lon = [name.split('_') for name in names]


    # Caricare il modello Streetview
    street_model = AutoModelForImageClassification.from_pretrained('sardegna-vit').to(device)
    street_classifier = pipeline("image-classification", model="sardegna-vit", device=device)

    # Inferenza sulle immagini Streetview
    predictions = [street_classifier(image) for image in street_valdala]
    predictions = [sard.convert_label(p[0]['label']) for p in predictions]

    walkability = pd.DataFrame(columns=['osmId','lat','lon','class'])
    for k in range(len(predictions)):
        walkability = walkability._append({'osmId': osm_lat_lon[k][0], 'lat': osm_lat_lon[k][1], 'lon': osm_lat_lon[k][2], 'class': predictions[k]}, ignore_index = True)
    walkability.to_csv(f'valdala_street_predictions.csv')

    # Estrarre le feature del modello Streetview
    if os.path.exists("valdala_features/street.npy"):
        street_features = np.load("valdala_features/street.npy")
    else:
        street_features = extract_inference_features(street_model, street_valdala, device, 'valdala_features/street.npy')

    # Parte 2 - Immagini Satellitari


    # Caricare le immagini Satellitari
    sat_valdala = sard.get_all_images_of_folder('C:/Users/mocci/Desktop/MOST/Val d\'Ala/satellite')

    # Caricare il modello Satellitare
    sat_model = AutoModelForImageClassification.from_pretrained('satellite-vit').to(device)
    sat_classifier = pipeline("image-classification", model="satellite-vit", device=device)

    # Inferenza sulle immagini Satellitari
    predictions = [sat_classifier(image) for image in sat_valdala]
    predictions = [sard.convert_label(p[0]['label']) for p in predictions]

    walkability = pd.DataFrame(columns=['lat','lon','class'])
    names = [os.path.splitext(os.path.basename(im.filename))[0] for im in sat_valdala]
    osm_lat_lon = [name.split('_') for name in names]

    for k in range(len(predictions)):
        walkability = walkability._append({'lat': osm_lat_lon[k][1], 'lon': osm_lat_lon[k][2], 'class': predictions[k]}, ignore_index = True)
    walkability.to_csv(f'valdala_sat_predictions.csv')

    # Estrarre le feature del modello Satellitare
    if os.path.exists("valdala_features/sat.npy"):
        sat_features = np.load("valdala_features/sat.npy")
    else:
        sat_features = extract_inference_features(street_model, street_valdala, device, 'valdala_features/sat.npy')

    # Parte 3 - Combinare i due modelli
    if os.path.exists('./valdala_features/street_sel.npy') and os.path.exists('./valdala_features/sat_sel.npy'):                                           # Se sono presenti le feature selezionate di training, caricale
        street_selected = np.load('valdala_features/street_sel.npy')
        sat_selected = np.load('valdala_features/sat_sel.npy')
    else:
        train_street_features = np.load('features/train_street_features.npy')
        train_sat_features = np.load('features/train_sat_features.npy')

        train_street, _, _ = sard.load_data()                                                                                                    # Carica i dataset per training, validation e test per le immagini street-view
        train_sat, _, _ = sard.load_data('satellite')                                                                                                  # Carica i dataset per training, validation e test per le immagini satellitari

        _, street_selector = fit_feature_selector(train_street_features,train_street['labels'],'features/train_street_selected.npy')           # altrimenti fitta un feature selector per street e sat sui dati di training e seleziona feature di train, valid e test
        _, sat_selector = fit_feature_selector(train_sat_features, train_sat['labels'],'features/train_sat_selected.npy')

        street_selected = select_features(street_selector, street_features, 'valdala_features/street_sel.npy')
        sat_selected = select_features(sat_selector, sat_features, 'valdala_features/sat_sel.npy')

   
    if os.path.exists("valdala_features/comb.npy"):                                                                                                              # Se sono presenti le feature concatenate di training
        combined = np.load("valdala_features/comb.npy")                                                                                                    # caricale per train, valid e test
    else:
        scaler = RobustScaler()                                                                                                                                    # normalizza feature                                                                                                                                   
        street_selected = scaler.fit_transform(street_selected)
        sat_selected = scaler.fit_transform(sat_selected)
        combined = concatenate_features(street_selected, sat_selected, "valdala_features/comb.npy")                                            # altrimenti concatena usando np.concatenate per train, valid e test

    predictions = inference_svm(combined)
    predictions = [p+1 for p in list(predictions)]

    for k in range(len(predictions)):
        walkability = walkability._append({'lat': osm_lat_lon[k][1], 'lon': osm_lat_lon[k][2], 'class': predictions[k]}, ignore_index = True)
        walkability.to_csv(f'comb_predictions.csv')

'''
Questa funzione addestra un modello dual encoder. Il primo encoder viene addestrato su immagini streetview, il secondo su
immagini satellitari e poi le feature estratte dagli encoder vengono concatenate e il modello da la sua predizione.
'''
def train_dual_encoder(checkpoint):


    batch_size = 16
    
    # Crea il dataset da utilizzare per l'addestramento: un'istanza di questo dataset rappresenta un punto (lat,lon)
    # con corrispondente immagine Streetview e Satellitare.
    dataset = StreetSatDataset()
    dataset.download_and_prepare()
    dataset_splits = dataset.as_dataset()
    
    # Split del dataset (80% train, 10% valid e 10% test)
    tf = paired.create_transforms(mode="both",checkpoint=checkpoint)
    train_split = dataset_splits["train"].with_transform(tf)
    val_split = dataset_splits["validation"].with_transform(tf)
    test_split = dataset_splits["test"].with_transform(tf)

    labels = train_split.unique('label')
    label2id = {c:idx for idx,c in enumerate(labels)}
    id2label = {idx:c for idx,c in enumerate(labels)}

    training_args = TrainingArguments(
        output_dir="./dual-encoder",
        eval_strategy="epoch",
        save_strategy="epoch",
        learning_rate=1e-4,
        per_device_train_batch_size=batch_size,
        per_device_eval_batch_size=batch_size,
        num_train_epochs=10,
        load_best_model_at_end=True,
        remove_unused_columns=False,
        fp16=True
    )
    model = de.DualEncoderModel(id2label=id2label, label2id=label2id, checkpoint = checkpoint)

    trainer = Trainer(
        model=model,
        args=training_args,
        train_dataset=train_split,
        eval_dataset=val_split,
        data_collator=paired.collate_fn,
        compute_metrics = sard.compute_metrics,
        tokenizer=AutoImageProcessor.from_pretrained(checkpoint, ignore_mismatched_sizes=True)
    )
    trainer.train()
    print("Validation Set")
    valid_res = trainer.evaluate()
    print("Test Set")
    test_res = trainer.evaluate(test_split)
    torch.save(model.state_dict(), './dual-encoder.pt')
    return valid_res, test_res




def save_checkpoint_to_excel(id, checkpoint, name, valid_sep, test_sep, valid_dual, test_dual, elapsed_sep, elapsed_dual):
        idx = id + 4                                # Le prime 3 righe sono intestazione della tabella
        workbook = load_workbook(filename=name)
        sheet = workbook.active

        sheet[f"A{idx}"] = checkpoint

        # Metriche Validation Set Esperimento Modelli Separati
        sheet[f"B{idx}"] = valid_sep['eval_accuracy']
        sheet[f"C{idx}"] = valid_sep['eval_precision']
        sheet[f"D{idx}"] = valid_sep['eval_recall']
        sheet[f"E{idx}"] = valid_sep['eval_f1']
        sheet[f"F{idx}"] = round(valid_sep['eval_loss'], 3)
        sheet[f"G{idx}"] = valid_sep['eval_one_out']

        # Metriche Test Set Esperimento Modelli Separati
        sheet[f"H{idx}"] = test_sep['eval_accuracy']
        sheet[f"I{idx}"] = test_sep['eval_precision']
        sheet[f"J{idx}"] = test_sep['eval_recall']
        sheet[f"K{idx}"] = test_sep['eval_f1']
        sheet[f"L{idx}"] = round(test_sep['eval_loss'], 3)
        sheet[f"M{idx}"] = test_sep['eval_one_out']
        sheet[f"N{idx}"] = round(elapsed_sep,1)
        sheet[f"O{idx}"] = test_sep['eval_mean_mae']

        # Metriche Validation Set Esperimento Modelli Uniti
        sheet[f"P{idx}"] = valid_dual['eval_accuracy']
        sheet[f"Q{idx}"] = valid_dual['eval_precision']
        sheet[f"R{idx}"] = valid_dual['eval_recall']
        sheet[f"S{idx}"] = valid_dual['eval_f1']
        sheet[f"T{idx}"] = round(valid_dual['eval_loss'], 3)
        sheet[f"U{idx}"] = valid_dual['eval_one_out']

        # Metriche Test Set Esperimento Modelli Uniti
        sheet[f"V{idx}"] = test_dual['eval_accuracy']
        sheet[f"W{idx}"] = test_dual['eval_precision']
        sheet[f"X{idx}"] = test_dual['eval_recall']
        sheet[f"Y{idx}"] = test_dual['eval_f1']
        sheet[f"Z{idx}"] = round(test_dual['eval_loss'], 3)
        sheet[f"AA{idx}"] = test_dual['eval_one_out']
        sheet[f"AB{idx}"] = round(elapsed_dual,1)
        sheet[f"AC{idx}"] = test_dual['eval_mean_mae']

        workbook.save(filename=name)

def model_testing_loop():
    checkpoint_names = ['google/vit-base-patch16-224', 'microsoft/swin-base-patch4-window7-224', 'facebook/deit-base-patch16-224',
                        'microsoft/beit-base-patch16-224', 'facebook/dinov2-base-imagenet1k-1-layer']

    for i in range(len(checkpoint_names)):

        if os.path.exists('features'):
            shutil.rmtree('features')
            os.makedirs('features')

        start_sep = time.time()
        valid_sep, test_sep = walkability_pipeline(checkpoint_names[i])
        end_sep = time.time()
        elapsed_sep = end_sep - start_sep
        sard.plot_confusion(test_sep['eval_confusion_matrix'],f"matrix/sep-{checkpoint_names[i].split('/')[-1]}")


        start_dual = time.time()
        valid_dual, test_dual = train_dual_encoder(checkpoint_names[i])
        end_dual = time.time()
        elapsed_dual = end_dual - start_dual
        sard.plot_confusion(test_dual['eval_confusion_matrix'],f"matrix/dual-{checkpoint_names[i].split('/')[-1]}")
        
        save_checkpoint_to_excel(id=i, checkpoint=checkpoint_names[i], name='transformers_experiment.xlsx', valid_sep=valid_sep, test_sep=test_sep, valid_dual=valid_dual, test_dual=test_dual, elapsed_sep=elapsed_sep, elapsed_dual=elapsed_dual)

def single_vit(checkpoint):

    batch_size = 16
    
    # Cancellare il dataset in cache in "C:\Users\mocci\.cache\huggingface\datasets"
    dataset = StreetSatDataset(mode="street")
    dataset.download_and_prepare()
    dataset_splits = dataset.as_dataset()
    
    # Split del dataset (80% train, 10% valid e 10% test)
    tf = paired.create_transforms(mode="street",checkpoint=checkpoint)
    train_split = dataset_splits["train"].with_transform(tf)
    val_split = dataset_splits["validation"].with_transform(tf)
    test_split = dataset_splits["test"].with_transform(tf)

    labels = train_split.unique('label')
    label2id = {c:idx for idx,c in enumerate(labels)}
    id2label = {idx:c for idx,c in enumerate(labels)}

    training_args = TrainingArguments(
        output_dir="./dual-encoder",
        evaluation_strategy="epoch",
        save_strategy="epoch",
        learning_rate=1e-4,
        per_device_train_batch_size=batch_size,
        per_device_eval_batch_size=batch_size,
        num_train_epochs=10,
        load_best_model_at_end=True,
        remove_unused_columns=False,
        fp16=True,
        optim='adamw_hf'
    )

    model = AutoModelForImageClassification.from_pretrained(checkpoint, label2id=label2id, id2label=id2label, ignore_mismatched_sizes=True)

    trainer = Trainer(
        model=model,
        args=training_args,
        train_dataset=train_split,
        eval_dataset=val_split,
        data_collator=paired.collate_fn,
        compute_metrics = sard.compute_metrics,
    )
    trainer.train()
    print("Validation Set")
    valid_res = trainer.evaluate()
    print("Test Set")
    test_res = trainer.evaluate(test_split)
    return valid_res, test_res

def single_testing_loop():
    dataset = StreetSatDataset(mode="both")
    dataset.download_and_prepare()
    dataset_splits = dataset.as_dataset()

    checkpoint_names = ['google/vit-base-patch16-224', 'microsoft/swin-base-patch4-window7-224', 'facebook/deit-base-patch16-224',
                        'microsoft/beit-base-patch16-224', 'facebook/dinov2-base-imagenet1k-1-layer']
    checkpoint_names = ['microsoft/swin-base-patch4-window7-224']

    for i in range(len(checkpoint_names)):
        name = 'new-dual-encoder.xlsx'
        idx = i + 3

        
        if os.path.exists('features'):
            shutil.rmtree('features')
            os.makedirs('features')
        
        start = time.time()
        # valid, test = walkability_pipeline(checkpoint_names[i])
        valid, test = single_vit(checkpoint_names[i])
        end = time.time()
        elapsed = end - start
        sard.plot_confusion(test['eval_confusion_matrix'],f"matrix/only-{checkpoint_names[i].split('/')[-1]}")
        workbook = load_workbook(filename = name)
        sheet = workbook.active

        sheet[f"A{idx}"] = checkpoint_names[i]

        # Metriche Validation Set Esperimento Modelli Separati
        sheet[f"B{idx}"] = test['eval_accuracy']
        sheet[f"C{idx}"] = test['eval_precision']
        sheet[f"D{idx}"] = test['eval_recall']
        sheet[f"E{idx}"] = test['eval_f1']
        sheet[f"F{idx}"] = round(test['eval_loss'], 3)
        sheet[f"G{idx}"] = test['eval_one_out']
        sheet[f"H{idx}"] = test['eval_mean_mae']
        sheet[f"I{idx}"] = elapsed


        workbook.save(filename = name)

def self_attention_fs(model, dataloader, filename):
    desired_percentile = 50
    use_threshold = True
    model.eval()
    all_attentions = []
    all_features = []
    device = torch.device('cuda')
    model.to(device)
    with torch.no_grad():
        for batch in tqdm(dataloader, desc=f"Processing output for {filename}"):
            inputs = batch['pixel_values'].to(device)
            outputs = model(inputs, output_hidden_states=True, output_attentions=True)
            feature = outputs.hidden_states[-1]
            all_features.append(feature.cpu())
            last_layer_attn = outputs.attentions[-1]

            cls_attn = last_layer_attn[:, :, 0, :]
            cls_attn = cls_attn.mean(dim=1)
            cls_attn = cls_attn[:, 1:]
            all_attentions.append(cls_attn.cpu())

            del inputs, outputs, feature, last_layer_attn, cls_attn
            torch.cuda.empty_cache()
    
    features = torch.cat(all_features, dim=0)
    patch_features = features[:, 1:, :]
    attentions = torch.cat(all_attentions, dim=0)

    seq_length = patch_features.size(1)
    top_k = max(1, round(0.5 * seq_length))
    global_attn = attentions.mean(dim=0)

    

    if use_threshold:
        threshold = torch.quantile(global_attn, q=desired_percentile / 100.0)
        selected_indices = (global_attn >= threshold).nonzero(as_tuple=True)[0]
        
        # If no patch meets the threshold, default to the patch with the highest global attention.
        if selected_indices.numel() == 0:
            selected_indices = torch.argmax(global_attn).unsqueeze(0)
    else:
        selected_indices = torch.argsort(global_attn, descending=True)[:top_k]
    
    global_top_indices = selected_indices.unsqueeze(0).expand(attentions.size(0), -1)
    # Expand to match the hidden dimension for gathering.
    # global_top_indices_expanded: [total_samples, top_k, hidden_size]
    global_top_indices_expanded = global_top_indices.unsqueeze(-1).expand(-1, -1, patch_features.size(-1))

    selected_features = torch.gather(patch_features, 1, global_top_indices_expanded)
    np.save(filename, selected_features.cpu().numpy())

    print("Original features shape:", patch_features.shape)
    print("Selected features shape:", selected_features.shape)
    return selected_features

if __name__ == '__main__':
    train_dual_encoder('microsoft/swin-base-patch4-window7-224')